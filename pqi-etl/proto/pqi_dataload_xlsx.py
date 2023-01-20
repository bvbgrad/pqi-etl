import csv
import re
import secrets

from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

BaseDir = 'pqi-etl/data/'
ChampionFile = BaseDir + '20230119 Master List of all SPEAK UP Champion Attendees.xlsx'
MemberFile = BaseDir + '20230119_member_export_ymx.xlsx'
NonMemberFile = BaseDir + '2023011901_non_member_export_ymx.xlsx'
UniqueNamesOutFile = BaseDir + 'out.member_unique_names'
NonMemberUniqueNamesOutFile = BaseDir + 'out.non_member_unique_names'
ChampionUniqueNamesOutFile = BaseDir + 'out.champion_unique_names'
ChampionEmailActionOutFile = BaseDir + 'out.champion_email_action.xlsx'
ChampionUpdateActionOutFile = BaseDir + 'out.champion_update_action.csv'
ChampionCreateActionOutFile = BaseDir + 'out.champion_create_action.csv'

RI = {'A':0, 'B':1, 'C':2, 'D':3, 'E':4, 'F':5, 'G':6, 'H':7, 'I':8, 'J':9, 
  'K':10, 'L':11, 'M':12, 'N':13, 'O':14, 'P':15, 'Q':16, 'R':17, 'S':18,
  'T':19, 'U':20, 'V':21, 'W':22, 'X':23, 'Y':24, 'Z':25}
NewMemberCols = {
  "Member Type Code": "",
  "Username": "",
  "Password": "",
  "First Name": "",
  "Last Name": "",
  "Email Address": "",
  "Middle Name": "",
  "Maiden Name": "",
  "Nickname": "",
  "Name Suffix": "",
  "Gender": "",
  # "Registration Date": "",
  "Member Approved": "",
  "Membership": "",
  # "Date Membership Expires": "",
  "Membership Expires": "",
  # "Date Last Renewed": "",
  "Email Bounced": "",
  "Home Address Line 1": "",
  "Home Address Line 2": "",
  "Home City": "",
  "Home Location": "",
  "Home Postal Code": "",
  "Home Country": "",
  "Home Phone Area Code": "",
  "Home Phone": "",
  "Mobile Area Code": "",
  "Mobile": "",
  "Organization Name": "",
  "Professional Title": "",
  "Profession": "",
  "Employer Address Line 1": "",
  "Employer Address Line 2": "",
  "Employer City": "",
  "Employer Location": "",
  "Employer Postal Code": "",
  "Employer Country": "",
  "Employer Phone Area Code": "",
  "Employer Phone": "",
  "Employer Website": "",
  "Internal Comments": "",
  "Champion": ""
}

usernameSet = set()
def generate_username(lastName, firstName):
  if lastName is None and firstName is None:
    userName = secrets.token_urlsafe(9)
  elif lastName == '' or lastName is None:
    userName = firstName[0:9]
  elif firstName == '' or firstName is None:
    userName = lastName[0:9]
  elif len(lastName) < 9:
    try:
      userName = lastName + firstName[0:1]
    except Exception as e:
      userName = 'no-name'
      print(e)
  else:
    userName = lastName[0:9] + firstName[0:1]
  userName = userName.lower()
  bduplicate = True
  numSuffix = 1
  # check if the username is unique by trying to add it to a set
  # assume not unique so will have at least one pass through the loop
  # if not unique, add monotonic incremented number until the new username is unique
  while bduplicate:
    try:
      if userName in usernameSet:
        userName = userName + str(numSuffix)
        numSuffix += 1
      else:
        usernameSet.add(userName)
        bduplicate = False
    except Exception as e:
      print(f"Error creating unique username: {e}")

  return userName


      # usernameSet.add(userName)
      # bduplicate = False

def generate_password():
  return secrets.token_urlsafe(8)


def find_download_data_row(dataRows, keyValues, emailColName):
  lastName, firstName, emailValue = keyValues
  dataRowFound = None

  for dataRow in dataRows:
    if dataRow[emailColName] == emailValue:
      dataRowFound = dataRow
  
  # email match is primary criteria, but if a match is not found
  # lastName & firstName are an alternate criteria
  if dataRowFound is None:
    for dataRow in dataRows:
      if dataRow['Last Name'] == lastName and dataRow['First Name'] == firstName:
        dataRowFound = dataRow

  return dataRowFound


def find_download_data_row_name (dataRows, lastName, firstName):
  for dataRow in dataRows:
    if dataRow['Last Name'] == lastName and dataRow['First Name'] == firstName:
      return dataRow


def fill_create_row_values(championRow, nonMemberRow):
  rowDict = NewMemberCols
  # Use alternate values for specific columns
  # Copy all other values from the non-member data
  for key in NewMemberCols.keys():
    if key == 'Champion':
      rowDict['Champion'] = 'Yes'
    elif key == 'Member Type Code':
      rowDict['Member Type Code'] = 'BasicI'
    elif key == 'Membership':
      rowDict['Membership'] = 'Basic Individual Subscription (Free)'
    elif key == 'Membership Expires':
      rowDict['Membership Expires'] = 'No'
    elif key == 'Member Approved':
      rowDict['Member Approved'] = 'Yes'
    elif key == 'Username':
      rowDict['Username'] = generate_username(championRow['Last Name'], championRow['First Name'])
    elif key == 'Password':
      rowDict['Password'] = generate_password()
    elif key == 'Organization Name':
      rowDict['Organization Name'] = championRow['Organization']
    else:
      rowDict[key] = nonMemberRow[key]

  return rowDict


def xlsx_reader(filename):
    rows = []
    try:
        wb = load_workbook(filename=filename, read_only=True)
        print(f"Load xlsx worksheet: {wb.sheetnames[0]}")
        ws = wb[wb.sheetnames[0]]
        bHeader = True
        headerRow = ()
        for row in ws.values:
          if bHeader:
            headerRow = row
            bHeader = False
            continue
          rowDict = dict(zip(headerRow, row))
          rows.append(rowDict)
        wb.close()
    except Exception as e:
        print(f"Supported formats are: .xlsx,.xlsm,.xltx,.xltm \n  {e}")
    wb.close()
    return rows


# rows are a dictionary
# columnKeys identify which columns to use for the sorted list tuple
def create_name_dataset_keyed(rows, columnKeys):
  lookupSet = set()
  lookupList = []
  print(f"Extracting data for columns: {columnKeys}")
  for i, row in enumerate(rows, 1):
    # Throw away the header row
    if i == 1: 
      continue
    rowTuple = ()
    try:
      for columnName in columnKeys:
        columnValue = row[columnName]
        columnValue = columnValue if columnValue is not None else ''
        rowTuple = (*rowTuple, columnValue,)
      lookupSet.add(rowTuple)
      lookupList.append(rowTuple)
    except Exception as e:
      print(f"Data error on row {i:5}. {e}")
      print(f"  {row}")
  sortedDataList = sorted(lookupList)
  print(f"Sorted data list: {len(sortedDataList)}")
  sortedDataset = sorted(lookupSet)
  print(f"Sorted data set: {len(sortedDataset)}")
  
  return sortedDataset


# rows are a tuple, columns specify which elements to use for the sorted list tuple
def create_name_dataset_indexed(rows, columnIndexes):
  lookupSet = set()
  lookupList = []
  print(f"Extracting data for columns: {columns}")
  for i, row in enumerate(rows, 1):
    # Throw away the header row
    if i == 1: 
      continue
    rowTuple = ()
    try:
      for columnName, columnIndex in columnIndexes:
        columnValue = row[columnIndex]
        columnValue = columnValue if columnValue is not None else ''
        rowTuple = (*rowTuple, columnValue,)
      lookupSet.add(rowTuple)
      lookupList.append(rowTuple)
    except Exception as e:
      print(f"Data error on row {i:5}. {e}")
      print(f"  {row}")
  sortedDataList = sorted(lookupList)
  print(f"Sorted data list: {len(sortedDataList)}")
  sortedDataset = sorted(lookupSet)
  print(f"Sorted data set: {len(sortedDataset)}")
  
  return sortedDataset

def create_csv(file_name_csv, dataRows):
  try:
    with open(file_name_csv, mode='w', newline='') as csv_out:
        writer = csv.writer(csv_out)
        writer.writerows(dataRows)
    print(f"CSV report saved to '{file_name_csv}'.  It has {len(dataRows)} rows.")
    print(f"Report saved at: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
  except Exception as err:
      print(f"Error saving csv worksheet\n  {err}")
      # sg.popup_error(f"Error saving account_summary report\n  {err}")


def create_xlsx_selected_columns(file_name, sheetName, columns, dataRows):
  columnNames = []
  for columnName in columns:
    columnNames.append(columnName)
  wb = Workbook()
  ws = wb.active
  ws.title = sheetName
  ws.append(columnNames)
  for row in dataRows:
      ws.append(row)
  ws.freeze_panes = 'A2'
  ws.auto_filter.ref = ws.dimensions
  ws["A1"].fill = PatternFill("solid", start_color="c9c9c9")

  try:
    file_name_xlsx = file_name + '.xlsx'
    wb.save(file_name_xlsx)
    print(f"Summary Excel report saved to '{file_name_xlsx}.xlsx'.  It has {len(dataRows)} rows.")

  except PermissionError:
      print(f"Could not save worksheet: '{file_name_xlsx}'\nCheck if a previous version is open in Excel.")
      # sg.popup_error("Could not save 'account_summary' report\nCheck if a previous version is open in Excel.")
      # window['-STATUS-'].update("'Save Report' operation canceled.")
  except Exception as err:
      print(f"Error saving worksheet\n  {err}")
      # sg.popup_error(f"Error saving account_summary report\n  {err}")
  finally:
      wb.close()


def adjust_name_columns(rowDict):
  newRowDict = rowDict
  if rowDict['Last Name'] == '' or rowDict['Last Name'] is None:
    # newFirstName = re.sub(r"[^a-zA-Z0-9 ]", "", rowDict['First Name'])
    newFirstName = rowDict['First Name']
    nameSplit = newFirstName.split(' ', 1)
    numNames = len(nameSplit)
    if numNames > 1:
      newRowDict['Last Name'] = nameSplit[1]
      newRowDict['First Name'] = nameSplit[0]
    else:
      newRowDict['Last Name'] = 'NLN'

    newRowDict['Last Name'] = newRowDict['Last Name'][0:1].upper() + newRowDict['Last Name'][1:]
    newRowDict['First Name'] = newRowDict['First Name'][0:1].upper() + newRowDict['First Name'][1:]
  return newRowDict


def create_champion_actions_xlsx(file_name, statusList, memberRows, nonMemberRows, championRows):
  upgradeMemberColumnNames = ['Last Name', 'First Name', 'Email', 'Username', 'Website ID', 'Champion']
  verifyColumnNames = ['Last Name', 'First Name', 'Email']
  
  wb = Workbook()
  ws = wb.active
  ws.title = 'Create'
  ws.append(list(NewMemberCols.keys()))
  for keyValues, status in statusList:
    if status == 'non-member':
      championRow = find_download_data_row(championRows, keyValues, 'Email')
      nonMemberRow = find_download_data_row(nonMemberRows, keyValues, 'Email Address') 
      if championRow is None and nonMemberRow is None:
        print(f"Champion and nonMember rows are empty: Skipping {keyValues}")
      elif championRow is None:
        print(f"Champion row empty: Skipping {keyValues}")
      elif nonMemberRow is None:
        print(f"nonMember row empty: Skipping {keyValues}")
      else:
        RowDict = fill_create_row_values(championRow, nonMemberRow)
        newRowDict = adjust_name_columns(RowDict)
        ws.append(list(newRowDict.values()))
        # ws.append(list(RowDict.values()))
    else:
      continue
  ws.freeze_panes = 'A2'
  ws.auto_filter.ref = ws.dimensions
  iStart = ord('A')
  if len(NewMemberCols) <= 26:
    iEnd = iStart + len(NewMemberCols)
  else:
    iEnd = iStart + 26
  for i in range(iStart, iEnd):
    ws[chr(i) + "1"].fill = PatternFill("solid", start_color="c9c9c9")

  wb.create_sheet('Upgrade')
  ws = wb['Upgrade']
  ws.append(upgradeMemberColumnNames)
  for keyValues, status in statusList:
    lastName, firstName, emailValue = keyValues
    if status == 'member':
      for item in memberRows:
        if item['Email Address'] == emailValue:
          memberID = item['Website ID (RO)'] ;
          # Column W in the member export spreadsheet
          memberUsername = item['Username'] ;
          championFlag = 'Yes' ;
          row = keyValues
          row = (*row, memberUsername, memberID, championFlag)
          ws.append(row)
    else:
      continue
  ws.freeze_panes = 'A2'
  ws.auto_filter.ref = ws.dimensions
  ws["A1"].fill = PatternFill("solid", start_color="c9c9c9")

  wb.create_sheet('Verify')
  ws = wb['Verify']
  ws.append(verifyColumnNames)
  for keyValues, status in statusList:
    if status == 'verify':
      row = keyValues
      ws.append(row)
    else:
      continue
  ws.freeze_panes = 'A2'
  ws.auto_filter.ref = ws.dimensions
  ws["A1"].fill = PatternFill("solid", start_color="c9c9c9")

  try:
    wb.save(file_name)
    print(f"Summary Excel report saved to '{file_name}'.")

  except PermissionError:
      print(f"Could not save worksheet: '{file_name}'\nCheck if a previous version is open in Excel.")
      # sg.popup_error("Could not save 'account_summary' report\nCheck if a previous version is open in Excel.")
      # window['-STATUS-'].update("'Save Report' operation canceled.")
  except Exception as err:
      print(f"Error saving worksheet\n  {err}")
      # sg.popup_error(f"Error saving account_summary report\n  {err}")
  finally:
      wb.close()


def find_champion_status(championNames, memberNames, nonMemberNames):
  statusList = []
  count = 0
  for champion in championNames:
    status = 'verify'
    lastNameChampion, firstNameChampion, emailChampion = champion
    emailChampion = emailChampion.lower()
    for keyValues in memberNames:
      lastName, firstName, emailMember = keyValues
      if emailChampion == emailMember.lower():
        status = 'member'
    for keyValues in nonMemberNames:
      lastName, firstName, emailNonMember = keyValues
      if emailChampion == emailNonMember.lower():
        status = 'non-member'
    if emailChampion is None or emailChampion == '':
      status = 'verify'
  # Email check failed: see if can find in YM downloads using First & Last names
    if status == 'verify':
      for keyValues in memberNames:
        lastName, firstName, emailMember = keyValues
        if (firstNameChampion == firstName) and (lastNameChampion == lastName):
          if emailChampion != emailMember:
            print(f"{firstNameChampion} {lastNameChampion} status changed based on name match:")
            print(f"   Champion email = '{emailChampion}' - Member email = '{emailMember}'")
          status = 'member'
          count += 1
          print(f"   Status changed from 'verify' to '{status}'.")
          break
      for keyValues in nonMemberNames:
        lastName, firstName, emailNonMember = keyValues
        if (firstNameChampion == firstName) and (lastNameChampion == lastName):
          if emailChampion != emailNonMember:
            print(f"{firstNameChampion} {lastNameChampion} status changed based on name match:")
            print(f"   Champion email = '{emailChampion}' - Non-member email = '{emailNonMember}'")
          status = 'non-member'
          count += 1
          print(f"   Status changed from 'verify' to '{status}'.")
  # Need champion key values in next stage of the process
    statusList.append((champion, status,))

  if count == 0:
    print("No status changes based on a name match process rather than an email match")
  else:
    print(f"There were {count} status changes based on a name match process rather than an email match")

  histogramCount = {}
  for champion, status in statusList:
    try:
        histogramCount[status] += 1
    except KeyError:
        histogramCount[status] = 1
  print(f"\nChampion status histogram = {histogramCount}")
  numChampionNames = len(championNames)
  numStatus = len(statusList)
  if numChampionNames == numStatus:
    print(f"All {numChampionNames} champion records have been processed.")
  else:
    print(f"Error: {numChampionNames} champion records does not match {numStatus} status determinations.")

  return statusList


# bvb TODO Add relative Path to the data folder
if __name__ == "__main__":

  memberRows = xlsx_reader(MemberFile)
  print(f"Members (xlsx): {len(memberRows)}")
  # Column Y (zero based index = 24) contains the email value in the member export spreadsheet
  # columns = [('Last Name', 3,), ('First Name', 4,), ('Email Address', 24,)]
  columns = ('Last Name', 'First Name', 'Email Address',)
  memberNames = create_name_dataset_keyed(memberRows, columns)
  print(f"Member unique names: {len(memberNames)}")
  create_xlsx_selected_columns(UniqueNamesOutFile, 'names', columns, memberNames)

  nonMemberRows = xlsx_reader(NonMemberFile)
  print(f"\nNon-Members (xlsx): {len(nonMemberRows)}")
  # Column U (zero based index = 20) contains the email value in the non_member export spreadsheet
  # columns = [('Last Name', 2,), ('First Name', 3,), ('Email Address', 20,)]
  columns = ('Last Name', 'First Name', 'Email Address',)
  nonMemberNames = create_name_dataset_keyed(nonMemberRows, columns)
  print(f"Non-Member unique names: {len(nonMemberNames)}")
  create_xlsx_selected_columns(NonMemberUniqueNamesOutFile, 'names', columns, nonMemberNames)

  championRows = xlsx_reader(ChampionFile)
  print(f"\nChampions (xlsx): {len(championRows)}")
  # Column C (zero based index = 2) contains the email value in the champion export spreadsheet
  # columns = [('Last Name', 1,), ('First Name', 0), ('Email', 2)]
  columns = ('Last Name', 'First Name', 'Email',)
  championNames = create_name_dataset_keyed(championRows, columns)
  print(f"Champion unique names: {len(championNames)}")
  create_xlsx_selected_columns(ChampionUniqueNamesOutFile, 'names', columns, championNames)

  statusList = find_champion_status(championNames, memberNames, nonMemberNames)
  create_champion_actions_xlsx(ChampionEmailActionOutFile, statusList, memberRows, nonMemberRows, championRows)

  upgradeCsvColumnNames = ['Website ID', 'Champion']
  dataRows = []
  dataRows.append(upgradeCsvColumnNames)
  for keyValues, status in statusList:
    lastName, firstName, emailValue = keyValues
    if status == 'member':
      for i, item in enumerate(memberRows, 1):
  # Column Y (email address)) in the member export spreadsheet
        if item['Email Address'] is not None:
          emailMember = item['Email Address'].lower()
          if emailMember == emailValue.lower():
    # Column L (Website ID (RO)) in the member export spreadsheet
            memberID = item['Website ID (RO)'] ;
            championFlag = 'Yes' ;
            row = list((memberID, championFlag,))
            dataRows.append(row)
        # else:
        #   print(f"working on {keyValues} Empty email on memberRow {i:5} ")
  create_csv(ChampionUpdateActionOutFile, dataRows)

  # WE DO USE CREATED FILE FOR INPUT
  dataRowsDict = xlsx_reader(ChampionEmailActionOutFile)
  dataRows = []
  bHeaderAdd = True
  for row in dataRowsDict:
    if bHeaderAdd:
      dataRows.append(row.keys())
      bHeaderAdd = False
    dataRows.append(list(row.values()))
  create_csv(ChampionCreateActionOutFile, dataRows)
